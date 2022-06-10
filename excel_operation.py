import glob
import os
# pip install openpyxl
from datetime import datetime

import openpyxl
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill

test_result_location = "test_result/Result.xlsx"
test_case_location = "test_case/test_case.xlsx"
test_result_location1 = "test_result/Result2.xlsx"


def clear_result():
    files = glob.glob('test_result/*')
    for file in files:
        os.remove(file)
    print("All existing test result has been removed")


def write_header():
    workbook = openpyxl.Workbook()  # This will create new workbook
    worksheet = workbook.create_sheet("Test_Result")  # This will add sheet in workbook
    worksheet.cell(1, 1, "SN")
    worksheet.cell(1, 2, "Test Summary")
    worksheet.cell(1, 3, "Result")
    worksheet.cell(1, 4, "Remarks")
    workbook.save(test_result_location)


def write_header1():
    workbook = openpyxl.load_workbook(test_case_location)
    worksheet = workbook.get_sheet_by_name("Sheet1")
    worksheet.cell(1, 6, "Result")
    worksheet.cell(1, 7, "Remarks")
    workbook.save(test_result_location1)


def write_result(sn, test_summary, result, remarks):
    workbook = openpyxl.load_workbook(test_result_location)
    worksheet = workbook.get_sheet_by_name("Test_Result")
    row = int(sn) + 1
    worksheet.cell(row, 1, sn)
    worksheet.cell(row, 2, test_summary)
    worksheet.cell(row, 3, result)
    worksheet.cell(row, 4, str(remarks))

    red_fill = PatternFill(start_color='EE1111', end_color='EE1111', fill_type='solid')
    green_fill = PatternFill(start_color='00AA00', end_color='00AA00', fill_type='solid')
    blue_fill = PatternFill(start_color='68A0F9', end_color='68A0F9', fill_type='solid')
    character = ('A', 'B', 'C', 'D')
    for ranges in character:
        cell = ranges + str(row)
        worksheet.conditional_formatting.add('A1:D1',
                                             FormulaRule(formula=['ISBLANK(L1)'], stopIfTrue=True, fill=blue_fill))
        worksheet.conditional_formatting.add(cell, FormulaRule(formula=['ISNUMBER(SEARCH("FAIL",' + cell + '))'],
                                                               stopIfTrue=True, fill=red_fill))
        worksheet.conditional_formatting.add(cell, FormulaRule(formula=['ISNUMBER(SEARCH("PASS",' + cell + '))'],
                                                               stopIfTrue=True, fill=green_fill))

    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter  # get column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 1)
        worksheet.column_dimensions[column].width = adjusted_width

    workbook.save(test_result_location)


def write_result2(sn, result, remarks):
    workbook = openpyxl.load_workbook(test_result_location1)
    worksheet = workbook.get_sheet_by_name("Sheet1")
    row = int(sn) + 1
    worksheet.cell(row, 6, result)
    worksheet.cell(row, 7, str(remarks))

    red_fill = PatternFill(start_color='EE1111', end_color='EE1111', fill_type='solid')
    green_fill = PatternFill(start_color='00AA00', end_color='00AA00', fill_type='solid')
    blue_fill = PatternFill(start_color='68A0F9', end_color='68A0F9', fill_type='solid')
    character = ('F', 'G')
    for ranges in character:
        cell = ranges + str(row)
        worksheet.conditional_formatting.add('A1:G1',
                                             FormulaRule(formula=['ISBLANK(L1)'], stopIfTrue=True, fill=blue_fill))
        worksheet.conditional_formatting.add(cell, FormulaRule(formula=['ISNUMBER(SEARCH("FAIL",' + cell + '))'],
                                                               stopIfTrue=True, fill=red_fill))
        worksheet.conditional_formatting.add(cell, FormulaRule(formula=['ISNUMBER(SEARCH("PASS",' + cell + '))'],
                                                               stopIfTrue=True, fill=green_fill))

    workbook.save(test_result_location1)


def writing_files(sn, test_summary, result, remarks):
    workbook, worksheet1, worksheet2 = excel_creater()
    filed_names = (int(sn), test_summary, result, str(remarks))
    start_column = 1
    start_row = int(sn) + 1
    for filed_names in filed_names:
        worksheet2.cell(row=start_row, column=start_column).value = filed_names
        start_column += 1
        workbook.save(test_result_location)


def write_summary():
    workbook = openpyxl.load_workbook(test_result_location)
    worksheet1 = workbook.create_sheet("Summary")
    worksheet1.cell(row=1, column=1).value = "Test Executed On:"
    worksheet1.cell(row=1, column=2 ).value = datetime.now()
    worksheet1.cell(row=2, column=1).value = "Total number of test"
    worksheet1.cell(row=2, column=2).value = "=COUNT(Test_Result!A:A)"
    worksheet1.cell(row=3, column=1).value = "Number of Passed test case"
    worksheet1.cell(row=3, column=2).value = '=COUNTIF(Test_Result!C:C,"PASS")'
    worksheet1.cell(row=4, column=1).value = "Number of failed test"
    worksheet1.cell(row=4, column=2).value = '=COUNTIF(Test_Result!C:C,"FAIL")'
    worksheet1.cell(row=5, column=1).value = "Number of Skipped tested case"
    worksheet1.cell(row=5, column=2).value = '=COUNTIF(Test_Result!D:D,"Test was skipped due to N flag")'
    fit_column(worksheet1)
    blue_fill = PatternFill(start_color='68A0F9', end_color='68A0F9', fill_type='solid')
    worksheet1.conditional_formatting.add('A1:A5', FormulaRule(formula=['ISBLANK(L1)'], stopIfTrue=True, fill=blue_fill))
    workbook.save(test_result_location)


# def format_excel():
#     workbook = openpyxl.load_workbook(test_result_location)
#     worksheet = workbook.get_sheet_by_name("Test_Result")
#
#     red_fill = PatternFill(start_color='EE1111', end_color='EE1111', fill_type='solid')
#     green_fill = PatternFill(start_color='00AA00', end_color='00AA00', fill_type='solid')
#     blue_fill = PatternFill(start_color='68A0F9', end_color='68A0F9', fill_type='solid')
#     character = ('A', 'B', 'C', 'D')
#     for ranges in character:
#         cell = ranges + str(row)
#         worksheet.conditional_formatting.add('A1:D1',
#                                              FormulaRule(formula=['ISBLANK(L1)'], stopIfTrue=True, fill=blue_fill))
#         worksheet.conditional_formatting.add(cell, FormulaRule(formula=['ISNUMBER(SEARCH("FAIL",' + cell + '))'],
#                                                                stopIfTrue=True, fill=red_fill))
#         worksheet.conditional_formatting.add(cell, FormulaRule(formula=['ISNUMBER(SEARCH("PASS",' + cell + '))'],
#                                                                stopIfTrue=True, fill=green_fill))


def fit_column(worksheet2):
    for col in worksheet2.columns:
        max_length = 0
        column = col[0].column_letter  # get column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 0.5)
        worksheet2.column_dimensions[column].width = adjusted_width

def excel_creater():
    fname = 'test_result/test_result.xlsx'
    if os.path.exists(fname):
        workbook = openpyxl.load_workbook(fname)
        worksheet1 = workbook.create_sheet('Summary')
        worksheet2 = workbook.create_sheet('Test_Result')
        return workbook, worksheet1, worksheet2
    else:
        workbook = openpyxl.Workbook()
        worksheet1 = workbook.create_sheet('Summary')
        worksheet2 = workbook.create_sheet('Test_Result')
        return workbook, worksheet1, worksheet2